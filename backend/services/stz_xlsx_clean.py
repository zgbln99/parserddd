"""
Clean already-generated Stundenzettel XLSX files — the spreadsheet counterpart
of ``stz_pdf_clean``.

Same edits as the PDF cleaner, on the source workbook:
  - clear the "Vorlage zur Dokumentation der täglichen Arbeitszeit" title,
  - drop the DATEV logo (openpyxl doesn't re-embed the picture on save),
  - remove the "Unterschrift des ..." labels; keep a centred "Datum" under the
    left signature line and put a centred "Kontrolle durch" under the right one,
  - set the print area to fit a single page.

Everything else (times, formulas, number formats, borders, fills, column
widths) is preserved by openpyxl.
"""

import io

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.properties import PageSetupProperties

_TITLE_MARK = "Vorlage zur Dokumentation"
_UNT_AN = "Unterschrift des Arbeitnehmers"
_UNT_AG = "Unterschrift des Arbeitgebers"


def _merge_end_col(ws, cell):
    """Right-most column of the merged range that starts at ``cell`` (or its own
    column if it isn't merged)."""
    for m in ws.merged_cells.ranges:
        if cell.coordinate in m:
            return m.max_col
    return cell.column


def _clean_sheet(ws) -> bool:
    title = None
    datums = []
    unt_an = unt_ag = None
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not isinstance(v, str):
                continue
            if _TITLE_MARK in v:
                title = c
            elif v.strip() == "Datum":
                datums.append(c)
            elif _UNT_AN in v:
                unt_an = c
            elif _UNT_AG in v:
                unt_ag = c

    changed = False
    if title is not None:
        title.value = None
        changed = True

    if len(datums) >= 2 and unt_an is not None and unt_ag is not None:
        datums.sort(key=lambda c: c.column)
        left_d, right_d = datums[0], datums[-1]
        row = left_d.row
        left_c0, left_c1 = left_d.column, _merge_end_col(ws, unt_an)
        right_c0, right_c1 = right_d.column, _merge_end_col(ws, unt_ag)

        # Unmerge any merged range that lives entirely in the footer row, so we
        # can re-merge each signature block cleanly.
        for m in [m for m in ws.merged_cells.ranges if m.min_row == row and m.max_row == row]:
            ws.unmerge_cells(str(m))

        for col in range(min(left_c0, right_c0), max(left_c1, right_c1) + 1):
            ws.cell(row=row, column=col).value = None

        def _label(col0, col1, text):
            cell = ws.cell(row=row, column=col0)
            cell.value = text
            cell.alignment = Alignment(horizontal="center", vertical=cell.alignment.vertical)
            if col1 > col0:
                ws.merge_cells(start_row=row, start_column=col0, end_row=row, end_column=col1)

        _label(left_c0, left_c1, "Datum")
        _label(right_c0, right_c1, "Kontrolle durch")
        changed = True

    # Fit to a single printed page.
    ws.page_setup.scale = None
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    else:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    return changed


def clean_xlsx(data: bytes) -> bytes:
    """Return cleaned XLSX bytes. Always re-serialises (openpyxl), so the DATEV
    logo is dropped even on an otherwise-already-clean file."""
    wb = openpyxl.load_workbook(io.BytesIO(data))
    for ws in wb.worksheets:
        _clean_sheet(ws)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
