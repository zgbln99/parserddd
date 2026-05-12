"""Fahrerliste — fill the accountant's monthly driver-list Excel from analyses.

Workflow:
  1. The accountant sends a monthly ``.xlsx`` with all drivers (one sheet per
     ~42-driver print page). The user uploads it here for a period (YYYY-MM).
  2. While analysing each driver, the frontend posts that driver's monthly
     numbers (per-day hours, 25 %, 40 %, VMA, AZ, Ur/Kr) to ``/fill``; we find
     the driver's row by name and write the cells, preserving everything else.
  3. The user downloads the filled workbook any time.

No data is recomputed here — the analysis layer provides the numbers; this
module only does name matching and cell writing on the uploaded file.
"""

from __future__ import annotations

import logging
import os
import re
import unicodedata
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file

from auth.decorators import login_required
from auth.helpers import _log_activity
from config import DATABASE_FILE

try:
    import openpyxl  # noqa: F401
    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    _HAS_OPENPYXL = False

_log = logging.getLogger(__name__)
bp = Blueprint('fahrerliste', __name__)

_STORE_DIR = os.path.join(os.path.dirname(DATABASE_FILE) or '.', 'fahrerliste')

# Day-cell markers we recognise / write straight through.
_ABSENCE_MARKERS = {'UR', 'KR', 'F', 'X'}

# Words inside a name cell that are notes, not part of the name.
_NAME_STOPWORDS = {
    'std', 'std.', 'hof', 'ross', 'a', 'n', 'b', 'h', 'b/h', 'dp', 'ro',
    'vma', 'gfb', 'tag', 'na', 'inkl', 'sa', 'so', 'incl', 'mit', 'ohne',
}


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _period_ok(period: str) -> bool:
    return bool(period) and bool(re.match(r'^\d{4}-\d{2}$', period))


def _store_path(period: str) -> str:
    return os.path.join(_STORE_DIR, f'{period}.xlsx')


def _ascii_fold(s: str) -> str:
    return ''.join(
        c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c)
    )


def _name_tokens(text) -> frozenset:
    """Token bag from a name cell — order- and case-independent.

    Takes the part before the first digit / ``%`` (the rest is notes),
    drops obvious note words, lowercases and strips diacritics.
    """
    if not text:
        return frozenset()
    s = str(text)
    s = re.split(r'[\d%]', s, 1)[0]
    s = _ascii_fold(s).lower()
    toks = [t for t in re.split(r'[^a-z]+', s) if len(t) >= 2 and t not in _NAME_STOPWORDS]
    return frozenset(toks)


def _hm_to_minutes(value) -> int | None:
    if value is None:
        return None
    m = re.match(r'^\s*(\d+):([0-5]?\d)\s*$', str(value))
    if not m:
        return None
    return int(m.group(1)) * 60 + int(m.group(2))


# ---------------------------------------------------------------------------
# workbook parsing
# ---------------------------------------------------------------------------

def _scan_sheet(ws):
    """Locate the day-header row and the special columns on one sheet.

    Returns ``None`` when the sheet doesn't look like a driver list, else::

        {
          "day_row": int,            # 1-based row with the day numbers
          "day_cols": {day:int -> col:int},
          "name_col": int,
          "n25_col"/"n40_col"/"ue_col"/"ur_col"/"kr_col"/"vma_col"/"az_col": int|None,
        }
    """
    max_r = min(ws.max_row, 6)
    max_c = ws.max_column
    best = None
    for r in range(1, max_r + 1):
        day_cols = {}
        markers = {}
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                iv = int(v)
                if 1 <= iv <= 31 and float(v) == iv:
                    day_cols[iv] = c
                elif abs(float(v) - 0.25) < 1e-6:
                    markers['n25_col'] = c
                elif abs(float(v) - 0.4) < 1e-6:
                    markers['n40_col'] = c
            elif isinstance(v, str):
                t = v.strip().lower()
                if t in ('ü', 'u', 'ue', 'überstunden'):
                    markers['ue_col'] = c
                elif t == 'ur':
                    markers['ur_col'] = c
                elif t == 'kr':
                    markers['kr_col'] = c
                elif t == 'vma':
                    markers['vma_col'] = c
                elif t == 'az':
                    markers['az_col'] = c
        if len(day_cols) >= 10 and best is None:
            best = (r, day_cols, markers)
    if not best:
        return None
    day_row, day_cols, markers = best

    # Name column: among A..(first day col - 1), the one with the most
    # name-ish strings in the rows below the header.
    first_day_col = min(day_cols.values())
    name_col, name_score = None, -1
    for c in range(1, first_day_col):
        score = 0
        for r in range(day_row + 1, min(ws.max_row, day_row + 60) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and len(v.strip()) > 4 and re.search(r'[A-Za-zÀ-ž]{3,}', v):
                score += 1
        if score > name_score:
            name_score, name_col = score, c
    if name_col is None or name_score <= 0:
        return None

    return {
        'day_row': day_row,
        'day_cols': day_cols,
        'name_col': name_col,
        'n25_col': markers.get('n25_col'),
        'n40_col': markers.get('n40_col'),
        'ue_col': markers.get('ue_col'),
        'ur_col': markers.get('ur_col'),
        'kr_col': markers.get('kr_col'),
        'vma_col': markers.get('vma_col'),
        'az_col': markers.get('az_col'),
    }


def _index_workbook(wb):
    """Return ``(sheets_meta, rows)`` where rows is a list of dicts:
    ``{sheet, row, name, tokens, meta}`` for every driver row found.
    """
    sheets_meta = {}
    rows = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        meta = _scan_sheet(ws)
        if not meta:
            continue
        sheets_meta[sn] = meta
        for r in range(meta['day_row'] + 1, ws.max_row + 1):
            nv = ws.cell(row=r, column=meta['name_col']).value
            if not isinstance(nv, str) or len(nv.strip()) <= 3:
                continue
            toks = _name_tokens(nv)
            if not toks:
                continue
            rows.append({'sheet': sn, 'row': r, 'name': nv.strip(), 'tokens': toks, 'meta': meta})
    return sheets_meta, rows


_PAGE_SIZE = 42


def _best_match(rows, driver_name):
    """Single best driver-row match among ``rows`` by name-token overlap.

    Returns ``(row_or_None, candidate_names)``.
    """
    want = _name_tokens(driver_name)
    if not want:
        return None, []
    scored = []
    for r in rows:
        common = want & r['tokens']
        if not common:
            continue
        scored.append((len(common) / len(want), r))
    if not scored:
        return None, []
    best_cov = max(c for c, _ in scored)
    if best_cov < 0.5:
        return None, [r['name'] for c, r in scored if c == best_cov][:5]
    top = [r for c, r in scored if abs(c - best_cov) < 1e-9]
    if len({r['tokens'] for r in top}) > 1:
        return None, [r['name'] for r in top][:5]
    return top[0], []


def _main_sheet_name(rows):
    """Sheet with the most driver rows = the 'index' sheet (e.g. LKW)."""
    counts = {}
    for r in rows:
        counts[r['sheet']] = counts.get(r['sheet'], 0) + 1
    return max(counts, key=counts.get) if counts else None


def _last_data_row(ws, meta):
    """Last row that still holds a driver name on a page/index sheet."""
    last = meta['day_row'] + 1
    for r in range(meta['day_row'] + 2, ws.max_row + 1):
        v = ws.cell(row=r, column=meta['name_col']).value
        if isinstance(v, str) and len(v.strip()) > 3:
            last = r
    return last


def _copy_row_style(ws, src_row, dst_row, max_col):
    from copy import copy as _copy
    for c in range(1, max_col + 1):
        s = ws.cell(row=src_row, column=c)
        if s.has_style:
            ws.cell(row=dst_row, column=c)._style = _copy(s._style)


def _find_sheet_ci(wb, name):
    """Case-insensitive lookup of a sheet by name; returns the actual name or None."""
    low = name.strip().lower()
    for sn in wb.sheetnames:
        if sn.strip().lower() == low:
            return sn
    return None


def _ensure_page_sheet(wb, name, template_sheet_name):
    """Return ``(actual_name, ws, meta)`` for page sheet ``name``; create it
    (header rows kept, driver rows cleared) from ``template_sheet_name`` when
    missing. Sheet lookup is case-insensitive.

    A freshly-cleared sheet has no name cells for :func:`_scan_sheet` to key
    on, so a new sheet reuses the template's scan metadata (column layout is
    identical after a copy).
    """
    existing = _find_sheet_ci(wb, name)
    if existing:
        ws = wb[existing]
        return existing, ws, _scan_sheet(ws)
    tpl_ws = wb[template_sheet_name]
    tpl_meta = _scan_sheet(tpl_ws)
    new = wb.copy_worksheet(tpl_ws)
    new.title = name
    if tpl_meta:
        first = tpl_meta['day_row'] + 2
        if new.max_row >= first:
            new.delete_rows(first, new.max_row - first + 1)
    return name, new, tpl_meta


# ---------------------------------------------------------------------------
# routes
# ---------------------------------------------------------------------------

@bp.route('/api/fahrerliste/upload', methods=['POST'])
@login_required
def fahrerliste_upload():
    if not _HAS_OPENPYXL:
        return jsonify({'error': 'openpyxl nie jest zainstalowany na serwerze'}), 500
    period = (request.form.get('period') or '').strip()
    if not _period_ok(period):
        return jsonify({'error': 'period (YYYY-MM) wymagany'}), 400
    if 'file' not in request.files:
        return jsonify({'error': 'Brak pliku'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'Wymagany plik .xlsx'}), 400
    os.makedirs(_STORE_DIR, exist_ok=True)
    path = _store_path(period)
    f.save(path)
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as exc:
        try:
            os.unlink(path)
        except OSError:
            pass
        return jsonify({'error': f'Nie można odczytać pliku: {exc}'}), 400
    _sheets_meta, rows = _index_workbook(wb)
    wb.close()
    _log_activity('fahrerliste_upload', f'{period} — {len(rows)} kierowców')
    return jsonify({
        'period': period,
        'count': len(rows),
        'drivers': sorted(r['name'] for r in rows),
        'sheets': sorted({r['sheet'] for r in rows}),
    })


@bp.route('/api/fahrerliste/status', methods=['GET'])
@login_required
def fahrerliste_status():
    if not _HAS_OPENPYXL:
        return jsonify({'exists': False, 'error': 'openpyxl nie jest zainstalowany'}), 200
    period = (request.args.get('period') or '').strip()
    if not _period_ok(period):
        return jsonify({'error': 'period (YYYY-MM) wymagany'}), 400
    path = _store_path(period)
    if not os.path.exists(path):
        return jsonify({'exists': False, 'period': period})
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        _m, rows = _index_workbook(wb)
        wb.close()
    except Exception as exc:
        return jsonify({'exists': True, 'period': period, 'error': str(exc)})
    return jsonify({
        'exists': True,
        'period': period,
        'count': len(rows),
        'drivers': sorted(r['name'] for r in rows),
        'sheets': sorted({r['sheet'] for r in rows}),
        'updated_at': datetime.utcfromtimestamp(os.path.getmtime(path)).isoformat() + 'Z',
    })


@bp.route('/api/fahrerliste/fill', methods=['POST'])
@login_required
def fahrerliste_fill():
    if not _HAS_OPENPYXL:
        return jsonify({'error': 'openpyxl nie jest zainstalowany na serwerze'}), 500
    data = request.get_json(silent=True) or {}
    period = (data.get('period') or '').strip()
    driver_name = (data.get('driver_name') or '').strip()
    if not _period_ok(period):
        return jsonify({'error': 'period (YYYY-MM) wymagany'}), 400
    if not driver_name:
        return jsonify({'error': 'driver_name wymagany'}), 400
    path = _store_path(period)
    if not os.path.exists(path):
        return jsonify({'error': f'Brak wgranej listy dla {period}'}), 404

    days = data.get('days') or {}            # {"1": "8:56", ...}
    absences = data.get('absences') or {}    # {"2": "Ur", ...}
    n25 = data.get('n25')                     # number or "X,XX" string
    n40 = data.get('n40')
    vma = data.get('vma')                     # int
    az = data.get('az')                       # "H:MM"
    ur = data.get('ur')                       # int (vacation days)
    kr = data.get('kr')                       # int (sick days)

    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as exc:
        return jsonify({'error': f'Nie można odczytać listy: {exc}'}), 500

    sheets_meta, rows = _index_workbook(wb)
    if not rows:
        wb.close()
        return jsonify({'error': 'Wgrana lista nie zawiera rozpoznawalnych arkuszy'}), 400

    main_sn = _main_sheet_name(rows)
    main_meta = sheets_meta[main_sn]
    main_ws = wb[main_sn]
    main_rows = [r for r in rows if r['sheet'] == main_sn]
    # The 'assignment' column on the index sheet: the one just left of the
    # name column (where the file already keeps 's4'/'s7'/...).
    assign_col = main_meta['name_col'] - 1 if main_meta['name_col'] > 1 else None

    # A page-style sheet to clone when a new s-sheet is needed.
    page_template = next((sn for sn in sheets_meta if sn != main_sn), main_sn)

    # ---- find / pick the target page sheet -------------------------------
    main_match, candidates = _best_match(main_rows, driver_name)
    display_name = main_match['name'] if main_match else driver_name

    assigned = None
    if main_match and assign_col:
        av = main_ws.cell(row=main_match['row'], column=assign_col).value
        if isinstance(av, str) and av.strip():
            assigned = av.strip()

    # how many drivers are currently assigned to each sheet (per index column),
    # keyed case-insensitively so "s1" and "S1" count together.
    assign_counts = {}
    if assign_col:
        for r in main_rows:
            av = main_ws.cell(row=r['row'], column=assign_col).value
            if isinstance(av, str) and av.strip():
                k = av.strip().lower()
                assign_counts[k] = assign_counts.get(k, 0) + 1

    if assigned:
        target_req = assigned
    else:
        target_req = 's1'
        i = 1
        while i < 100:
            name = f's{i}'
            if assign_counts.get(name.lower(), 0) < _PAGE_SIZE:
                target_req = name
                break
            i += 1

    # never write the index sheet itself as a "page"
    if (_find_sheet_ci(wb, target_req) or '').strip() == main_sn or target_req.lower() == main_sn.lower():
        target_req = 's_pages'

    target_sn, target_ws, target_meta = _ensure_page_sheet(wb, target_req, page_template)
    if target_meta is None:
        wb.close()
        return jsonify({'error': f'Nie udało się przygotować arkusza {target_req}'}), 500

    def _num_str(v):
        if v is None or v == '':
            return None
        if isinstance(v, str):
            return v.replace('.', ',')
        return f'{float(v):.2f}'.replace('.', ',')

    def _fill_row(ws, r, meta):
        for col in meta['day_cols'].values():
            ws.cell(row=r, column=col, value=None)
        for d, col in meta['day_cols'].items():
            d_str = str(d)
            marker = absences.get(d_str)
            if marker and str(marker).strip().upper() in _ABSENCE_MARKERS:
                ws.cell(row=r, column=col, value=str(marker).strip())
                continue
            val = days.get(d_str)
            if val not in (None, ''):
                ws.cell(row=r, column=col, value=str(val))

        def _set(col, value):
            if col and value is not None and value != '':
                ws.cell(row=r, column=col, value=value)

        _set(meta.get('n25_col'), _num_str(n25))
        _set(meta.get('n40_col'), _num_str(n40))
        if meta.get('vma_col') and vma not in (None, ''):
            try:
                ws.cell(row=r, column=meta['vma_col'], value=int(vma))
            except (TypeError, ValueError):
                ws.cell(row=r, column=meta['vma_col'], value=vma)
        _set(meta.get('az_col'), str(az) if az not in (None, '') else None)
        for key, v in (('ur_col', ur), ('kr_col', kr)):
            col = meta.get(key)
            if col and v not in (None, '', 0):
                try:
                    ws.cell(row=r, column=col, value=int(v))
                except (TypeError, ValueError):
                    pass

    filled = []

    # ---- index sheet: fill data + write the page-sheet name --------------
    if main_match:
        main_r = main_match['row']
    else:
        main_r = _last_data_row(main_ws, main_meta) + 1
        _copy_row_style(main_ws, main_r - 1, main_r, main_ws.max_column)
        main_ws.cell(row=main_r, column=main_meta['name_col'], value=display_name)
    _fill_row(main_ws, main_r, main_meta)
    if assign_col:
        main_ws.cell(row=main_r, column=assign_col, value=target_sn)
    filled.append({'sheet': main_sn, 'row': main_r})

    # ---- target page sheet: update existing row or append a new one ------
    page_rows = []
    tlast = target_meta['day_row'] + 1
    for r in range(target_meta['day_row'] + 2, target_ws.max_row + 1):
        nv = target_ws.cell(row=r, column=target_meta['name_col']).value
        if isinstance(nv, str) and len(nv.strip()) > 3:
            tlast = r
            toks = _name_tokens(nv)
            if toks:
                page_rows.append({'sheet': target_sn, 'row': r, 'name': nv.strip(),
                                  'tokens': toks, 'meta': target_meta})
    pm, _cands = _best_match(page_rows, driver_name)
    if pm:
        page_r = pm['row']
    else:
        page_r = tlast + 1
        _copy_row_style(target_ws, page_r - 1, page_r,
                        max(target_meta['day_cols'].values()) if target_meta['day_cols'] else target_ws.max_column)
        target_ws.cell(row=page_r, column=target_meta['name_col'], value=display_name)
    _fill_row(target_ws, page_r, target_meta)
    filled.append({'sheet': target_sn, 'row': page_r})

    try:
        wb.save(path)
    except Exception as exc:
        wb.close()
        return jsonify({'error': f'Nie można zapisać listy: {exc}'}), 500
    wb.close()
    _log_activity('fahrerliste_fill', f'{period} — {driver_name} → {main_sn}/{target_sn}')
    return jsonify({
        'ok': True,
        'matched_name': display_name,
        'index_sheet': main_sn,
        'target_sheet': target_sn,
        'filled': filled,
        'created_sheet': target_sn if target_sn not in sheets_meta else None,
        'candidates': candidates if not main_match else [],
    })


@bp.route('/api/fahrerliste/download', methods=['GET'])
@login_required
def fahrerliste_download():
    period = (request.args.get('period') or '').strip()
    if not _period_ok(period):
        return jsonify({'error': 'period (YYYY-MM) wymagany'}), 400
    path = _store_path(period)
    if not os.path.exists(path):
        return jsonify({'error': f'Brak wgranej listy dla {period}'}), 404
    return send_file(
        path, as_attachment=True,
        download_name=f'fahrerliste_{period}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
